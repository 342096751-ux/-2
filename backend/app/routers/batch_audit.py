from __future__ import annotations

import io
from datetime import datetime
from fastapi import APIRouter, File, HTTPException, Query, UploadFile
from fastapi.responses import StreamingResponse
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from app.core.aggregator import STRATEGY_NAMES
from app.core.batch_audit_manager import BatchAuditManager


def parse_excel_content(file_content: bytes) -> list[dict[str, str]]:
    """解析 Excel，读取 id + content 两列"""
    wb = load_workbook(io.BytesIO(file_content), data_only=True, read_only=True)
    ws = wb.active

    headers: list[str] = []
    for cell in next(ws.iter_rows(min_row=1, max_row=1, values_only=True)):
        headers.append(str(cell).strip().lower() if cell is not None else "")

    id_idx: int | None = None
    content_idx: int | None = None

    for i, h in enumerate(headers):
        if h in ["id", "编号", "序号", "no"]:
            id_idx = i
        if h in ["content", "内容", "文本", "text"]:
            content_idx = i

    if content_idx is None:
        raise HTTPException(status_code=400, detail="Excel 中未找到 'content' 或 '内容' 列")

    rows: list[dict[str, str]] = []
    for row_data in ws.iter_rows(min_row=2, values_only=True):
        if row_data is None:
            continue
        row_tuple = tuple(row_data)
        row_id = ""
        if id_idx is not None and id_idx < len(row_tuple) and row_tuple[id_idx] is not None:
            row_id = str(row_tuple[id_idx]).strip()
        content = ""
        if content_idx < len(row_tuple) and row_tuple[content_idx] is not None:
            content = str(row_tuple[content_idx]).strip()
        if content:
            rows.append(
                {"id": row_id or f"row-{len(rows)+1}", "content": content}
            )
    return rows


def create_batch_audit_router(batch_manager: BatchAuditManager) -> APIRouter:
    router = APIRouter(prefix="/api/batch-audit", tags=["batch-audit"])

    allowed_strategies = set(STRATEGY_NAMES.keys())

    @router.post("/upload")
    async def create_batch_audit(
        file: UploadFile = File(...),
        concurrency: int = Query(3, ge=1, le=10, description="并发数，1-10"),
        strategy: str = Query("one_vote_veto", description="聚合策略"),
    ):
        """
        创建批量审核任务

        - 上传 Excel (.xlsx)，包含 id + content 两列
        - concurrency: 同时审核数量（默认 3）
        - strategy: 聚合策略（默认一票否决）
        """
        if strategy not in allowed_strategies:
            raise HTTPException(
                status_code=400,
                detail=f"不支持的 strategy，可选: {', '.join(sorted(allowed_strategies))}",
            )
        raw = await file.read()
        if len(raw) == 0:
            raise HTTPException(status_code=400, detail="文件为空")
        if len(raw) > 10 * 1024 * 1024:
            raise HTTPException(status_code=413, detail="文件超过 10MB 限制")

        try:
            rows = parse_excel_content(raw)
        except HTTPException:
            raise
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"解析 Excel 失败: {e}") from e

        total = len(rows)
        if total == 0:
            raise HTTPException(status_code=400, detail="未解析出有效数据")
        if total > 5000:
            raise HTTPException(
                status_code=400,
                detail=f"条目过多 ({total})，单次上限 5000 条",
            )

        task_id = await batch_manager.create_and_start(rows, concurrency, strategy)
        return {
            "task_id": task_id,
            "total": total,
            "concurrency": concurrency,
            "strategy": strategy,
            "message": f"任务已创建，共 {total} 条待审",
            "status_url": f"/api/batch-audit/{task_id}/status",
            "results_url": f"/api/batch-audit/{task_id}/results",
        }

    @router.get("/{task_id}/status")
    async def get_task_status(task_id: str):
        status = batch_manager.get_status(task_id)
        if not status:
            raise HTTPException(status_code=404, detail="任务不存在或已过期")
        return status

    @router.get("/{task_id}/results")
    async def get_task_results(
        task_id: str,
        limit: int = Query(50, ge=1, le=200),
        offset: int = Query(0, ge=0),
    ):
        """获取批量审核结果（分页）"""
        results = batch_manager.get_results(task_id)
        if results is None:
            raise HTTPException(status_code=404, detail="任务不存在")

        total = len(results)
        items = results[offset : offset + limit]

        return {
            "total": total,
            "completed": sum(1 for r in results if r.get("status") == "completed"),
            "failed": sum(1 for r in results if r.get("status") == "failed"),
            "items": items,
        }

    @router.get("/{task_id}/export")
    async def export_batch_results(task_id: str):
        """导出批量审核结果为 Excel 宽表"""
        results = batch_manager.get_results(task_id)
        if results is None:
            raise HTTPException(status_code=404, detail="任务不存在")
        if not results:
            raise HTTPException(status_code=404, detail="无结果")

        wb = Workbook()
        ws = wb.active
        assert ws is not None
        ws.title = "批量审核结果"

        column_widths = {
            "A": 20,
            "B": 40,
            "C": 12,
            "D": 10,
            "E": 30,
            "F": 12,
            "G": 10,
            "H": 12,
            "I": 10,
            "J": 12,
            "K": 10,
            "L": 12,
            "M": 30,
        }
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width

        headers = [
            "ID",
            "待审内容",
            "规则执行员",
            "规则置信度",
            "规则理由",
            "对抗侦探",
            "判例执行员",
            "判例置信度",
            "大法官",
            "大法官置信度",
            "最终结果",
            "最终置信度",
            "异常信息",
        ]

        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        header_fill = PatternFill(
            start_color="44546A", end_color="44546A", fill_type="solid"
        )
        header_font = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")

        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border

        verdict_fill = {
            "违规": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
            "正常": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
            "疑似": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
            "不参与": PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"),
            "未介入": PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"),
        }

        for row_idx, result in enumerate(results, 2):
            rule = result.get("rule_executor") or {}
            adv = result.get("adversarial_detective") or {}
            case = result.get("case_executor") or {}
            judge = result.get("chief_judge") or {}

            row_data = [
                result.get("id", ""),
                (result.get("content") or "")[:500],
                rule.get("verdict", "-"),
                rule.get("confidence", "-"),
                (rule.get("reason", "") or "")[:200],
                adv.get("verdict", "-"),
                case.get("verdict", "-"),
                case.get("confidence", "-"),
                judge.get("verdict", "-"),
                judge.get("confidence", "-"),
                result.get("final_result", "-"),
                result.get("final_confidence", "-"),
                result.get("error") or "",
            ]

            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.font = Font(name="微软雅黑", size=10, color="000000")
                cell.alignment = Alignment(
                    horizontal="left", vertical="center", wrap_text=True
                )
                cell.border = thin_border

                if col_idx in {3, 6, 7, 9, 11} and value in verdict_fill:
                    cell.fill = verdict_fill[value]

        ws.freeze_panes = "A2"

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        blob = buf.getvalue()

        filename = f"批量审核结果_{datetime.now():%Y%m%d_%H%M%S}.xlsx"

        return StreamingResponse(
            io.BytesIO(blob),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f'attachment; filename="{filename}"',
                "Content-Length": str(len(blob)),
            },
        )

    return router
