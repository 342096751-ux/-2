"""Excel export for audit trails (wide table)."""

from __future__ import annotations

import io
import json
from datetime import datetime
from typing import Any
from urllib.parse import quote

from fastapi import APIRouter, HTTPException, Query
from fastapi.responses import StreamingResponse
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from app.core.blackboard import Blackboard

# Excel 单元格上限约 32767 字符（留余量）
_EXCEL_CELL_MAX = 32700


def _format_llm_dialogue(trace: Any) -> str:
    """将黑板中的 llm_dialogue/trace 转成可读的整块文本。"""
    if trace is None or trace == {} or trace == []:
        return "-"
    if isinstance(trace, dict):
        lines: list[str] = []
        if trace.get("model"):
            lines.append(f"[模型] {trace['model']}")
        if trace.get("temperature") is not None:
            lines.append(f"[temperature] {trace['temperature']}")
        if trace.get("error"):
            lines.append(f"[错误] {trace['error']}")
        sp = trace.get("system_prompt")
        if isinstance(sp, str) and sp.strip():
            lines.append("[system_prompt]\n" + sp.strip())
        up = trace.get("user_prompt")
        if isinstance(up, str) and up.strip():
            lines.append("[user_prompt]\n" + up.strip())
        rsp = trace.get("response")
        if rsp is not None and str(rsp).strip():
            lines.append("[模型回复/raw]\n" + str(rsp).strip())
        if not lines:
            return json.dumps(trace, ensure_ascii=False)[:_EXCEL_CELL_MAX]
        out = "\n\n".join(lines)
        return out[:_EXCEL_CELL_MAX] + ("\n...(已截断)" if len(out) > _EXCEL_CELL_MAX else "")
    return json.dumps(trace, ensure_ascii=False)[:_EXCEL_CELL_MAX]


def create_export_router(bb: Blackboard) -> APIRouter:
    router = APIRouter(prefix="/api/export", tags=["export"])

    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    batch_header_fill = PatternFill(start_color="44546A", end_color="44546A", fill_type="solid")
    batch_header_font = Font(size=14, bold=True, color="FFFFFF")

    info_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    info_font = Font(size=11, color="333333")

    agent_colors = {
        "规则执行员": "C6EFCE",
        "对抗侦探": "FFEB9C",
        "判例执行员": "B8CCE4",
        "大法官": "FFC7CE",
        "聚合器": "E4DFEC",
        "系统": "D9E1F2",
        "default": "D9D9D9",
    }

    header_font = Font(size=11, bold=True)
    cell_font = Font(size=10)

    verdict_fill = {
        "违规": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
        "正常": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
        "疑似": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
        "不参与": PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"),
        "未介入": PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"),
    }

    verdict_font_cn = {
        "违规": Font(size=11, bold=True, color="9C0006"),
        "正常": Font(size=11, bold=True, color="006100"),
    }
    neutral_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

    agent_name_map = {
        "rule_executor": "规则执行员",
        "adversarial_detective": "对抗侦探",
        "case_executor": "判例执行员",
        "chief_judge": "大法官",
        "aggregator": "聚合器",
        "system": "系统",
        "unknown": "未知",
    }

    zone_map = {
        "rule_zone": "规则区",
        "adversarial_zone": "对抗区",
        "case_zone": "判例区",
        "final_zone": "终审区",
        "aggregate_zone": "聚合区",
        "orchestrator": "编排",
        "knowledge_zone": "知识区",
    }

    verdict_map = {
        "violation": "违规",
        "normal": "正常",
        "uncertain": "疑似",
        "not_participating": "不参与",
        "not_intervene": "未介入",
        "not_participate": "不参与",
        "running": "-",
        "": "-",
    }

    def zh_agent(agent_en: str) -> str:
        return agent_name_map.get(agent_en, agent_en)

    def zh_zone(zone_en: str) -> str:
        return zone_map.get(zone_en, zone_en or "-")

    def zh_verdict(verdict_raw: Any) -> str:
        if verdict_raw is None or verdict_raw == "":
            return "-"
        s = str(verdict_raw).strip().lower()
        return verdict_map.get(s, verdict_map.get(str(verdict_raw), str(verdict_raw)))

    def build_wide_table_excel(audit_ids: list[str]) -> bytes:
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        ws.title = "审核宽表"
        last_col = 9

        col_widths = [15, 12, 8, 18, 12, 10, 36, 45, 70]
        for idx, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(idx)].width = w

        current_row = 1

        for audit_idx, audit_id in enumerate(audit_ids):
            logs = bb.get_logs(audit_id)
            if not logs:
                continue

            meta_content = bb.get_state(audit_id, "content", "") or ""
            final_result = bb.get_state(audit_id, "final_verdict", "") or ""

            ts = logs[0].timestamp if logs else ""
            if not meta_content:
                meta_content = f"(未保存正文 audit_id={audit_id})"

            ws.merge_cells(
                start_row=current_row, start_column=1, end_row=current_row, end_column=last_col
            )
            title_cell = ws.cell(row=current_row, column=1)
            title_cell.value = f"审核批次: {audit_id}    时间: {ts}"
            title_cell.fill = batch_header_fill
            title_cell.font = batch_header_font
            title_cell.alignment = Alignment(horizontal="left", vertical="center")
            title_cell.border = thin_border
            ws.row_dimensions[current_row].height = 28
            current_row += 1

            ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=2)
            ws.cell(row=current_row, column=1).value = "输入内容:"
            ws.cell(row=current_row, column=1).font = Font(size=11, bold=True)
            ws.cell(row=current_row, column=1).fill = info_fill
            ws.cell(row=current_row, column=1).border = thin_border

            ws.merge_cells(start_row=current_row, start_column=3, end_row=current_row, end_column=last_col)
            c_cell = ws.cell(row=current_row, column=3)
            c_cell.value = meta_content[: _EXCEL_CELL_MAX]
            c_cell.font = info_font
            c_cell.fill = info_fill
            c_cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            c_cell.border = thin_border
            ws.row_dimensions[current_row].height = max(22, min(120, len(meta_content) // 30 + 20))
            current_row += 1

            ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=2)
            ws.cell(row=current_row, column=1).value = "最终结果:"
            ws.cell(row=current_row, column=1).font = Font(size=11, bold=True)
            ws.cell(row=current_row, column=1).fill = info_fill
            ws.cell(row=current_row, column=1).border = thin_border

            fv_cn = zh_verdict(final_result)
            ws.merge_cells(start_row=current_row, start_column=3, end_row=current_row, end_column=last_col)
            merged_cell = ws.cell(row=current_row, column=3)
            merged_cell.value = fv_cn
            merged_cell.font = verdict_font_cn.get(fv_cn, Font(size=11, bold=True))
            merged_cell.fill = verdict_fill.get(fv_cn, neutral_fill)
            merged_cell.alignment = Alignment(horizontal="left", vertical="center")
            merged_cell.border = thin_border
            ws.row_dimensions[current_row].height = 24
            current_row += 1
            current_row += 1

            headers_cn = [
                "Agent名称",
                "区域",
                "序号",
                "阶段",
                "判定",
                "置信度",
                "理由",
                "原始输出",
                "模型对话/轨迹",
            ]
            for col_idx, header in enumerate(headers_cn, 1):
                cell = ws.cell(row=current_row, column=col_idx, value=header)
                cell.font = header_font
                cell.fill = PatternFill(start_color="B8CCE4", end_color="B8CCE4", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = thin_border
            ws.row_dimensions[current_row].height = 22
            current_row += 1

            raw_entries = [e.model_dump(mode="json") for e in logs]
            grouped: dict[str, list[dict[str, Any]]] = {}
            for ent in raw_entries:
                ae = ent.get("agent", "unknown")
                grouped.setdefault(ae if isinstance(ae, str) else "unknown", []).append(ent)

            agent_order = ["rule_enforcer", "adversarial_detective", "case_executor", "confidence_evaluator", "chief_judge", "aggregator", "system", "unknown"]
            sorted_agent_keys = [k for k in agent_order if k in grouped]
            extra_keys = [k for k in grouped.keys() if k not in agent_order]

            hidden_phases = {"phase_change"}
            phase_order = {
                "variant_pre_check": 1,
                "rule_enforcer": 2,
                "knowledge_retriever": 3,
                "adversarial_detective": 4,
                "case_executor": 5,
                "confidence_assessor": 6,
                "chief_judge": 7,
                "aggregator": 8,
                "started": 0,
                "completed": 99,
            }

            for agent_en in sorted_agent_keys + sorted(extra_keys):
                agent_logs = [
                    entry for entry in grouped[agent_en]
                    if str(entry.get("phase", "") or "") not in hidden_phases
                ]
                if not agent_logs:
                    continue
                agent_logs = sorted(
                    agent_logs,
                    key=lambda x: (
                        phase_order.get(str(x.get("phase", "") or ""), 50),
                        str(x.get("timestamp", "")),
                    ),
                )
                agent_cn = zh_agent(agent_en)
                zone_cn = zh_zone(str(agent_logs[0].get("zone", "") or ""))
                shade = agent_colors.get(agent_cn, agent_colors["default"])

                for step_idx, entry in enumerate(agent_logs, 1):
                    data = entry.get("data") or {}
                    vd = ""
                    if isinstance(data, dict):
                        vd = data.get("verdict", "")
                        if vd == "" and "need_knowledge" in data:
                            vd = "-"
                    verdict_cn = zh_verdict(vd)
                    phase_cn = str(entry.get("phase", "") or "-")
                    raw_out = entry.get("raw_llm_output") or ""
                    reason = ""
                    llm_diag_text = "-"
                    if isinstance(data, dict):
                        reason = str(data.get("reason", "") or "")
                        diag = data.get("llm_dialogue")
                        if diag is not None:
                            llm_diag_text = _format_llm_dialogue(diag)
                    row_values = [
                        agent_cn,
                        zone_cn,
                        str(step_idx),
                        phase_cn,
                        verdict_cn,
                        str(data.get("confidence", "-")) if isinstance(data, dict) else "-",
                        reason[:1600],
                        (raw_out[:8000]) if isinstance(raw_out, str) else "",
                        llm_diag_text,
                    ]

                    for col_idx, value in enumerate(row_values, 1):
                        cell = ws.cell(row=current_row, column=col_idx, value=value)
                        cell.font = cell_font
                        cell.alignment = Alignment(
                            horizontal="left" if col_idx in {7, 8, 9} else "center",
                            vertical="center",
                            wrap_text=True,
                        )
                        cell.border = thin_border

                        if col_idx == 1:
                            cell.fill = PatternFill(
                                start_color=shade,
                                end_color=shade,
                                fill_type="solid",
                            )

                        if col_idx == 5 and verdict_cn in verdict_fill:
                            cell.fill = verdict_fill[verdict_cn]

                    dh = len(str(llm_diag_text)) // 50 + len(str(reason)) // 45 + len(str(raw_out)) // 50
                    ws.row_dimensions[current_row].height = min(409, max(48, dh * 4))
                    current_row += 1

            current_row += 2

            if audit_idx < len(audit_ids) - 1:
                ws.merge_cells(
                    start_row=current_row, start_column=1, end_row=current_row, end_column=last_col
                )
                sep = ws.cell(row=current_row, column=1)
                sep.border = Border(
                    top=Side(style="medium", color="44546A"),
                    bottom=Side(style="medium", color="44546A"),
                )
                ws.row_dimensions[current_row].height = 8
                current_row += 1

        ws.freeze_panes = "A2"

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf.getvalue()

    @router.get("/wide-table/excel")
    async def export_wide_table_excel(
        audit_ids: list[str] | None = Query(default=None),
        sort: str = Query(default="desc"),
    ):
        aids = audit_ids if audit_ids else bb.list_audit_ids_by_recency(limit=10)
        aids = list(dict.fromkeys(aids))
        aids = [a for a in aids if bb.get_logs(a)]
        if sort.lower() == "asc":
            aids = list(reversed(aids))
        if not aids:
            raise HTTPException(status_code=404, detail="未找到任何审核记录")
        excel_bytes = build_wide_table_excel(aids)
        ts = datetime.now()
        name_cn = f"审核宽表_{ts:%Y%m%d_%H%M%S}.xlsx"
        name_ascii = f"audit_wide_{ts:%Y%m%d_%H%M%S}.xlsx"
        content_disp = (
            f'attachment; filename="{name_ascii}"; '
            f"filename*=UTF-8''{quote(name_cn)}"
        )
        return StreamingResponse(
            io.BytesIO(excel_bytes),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": content_disp},
        )

    @router.get("/history/list")
    async def audit_history_list(
        limit: int = Query(default=50, ge=1, le=200),
        offset: int = Query(default=0, ge=0),
    ) -> list[dict[str, Any]]:
        aids = bb.list_audit_ids_by_recency(limit=limit + offset, offset=0)
        slice_ids = aids[offset : offset + limit]
        out: list[dict[str, Any]] = []
        for aid in slice_ids:
            out.append(bb.audit_summary(aid))
        return out

    return router
